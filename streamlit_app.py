import streamlit as st
import pandas as pd
import tempfile
import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 초기 상태 설정
if 'username' not in st.session_state:
    st.session_state['username'] = ''
if 'time_unit' not in st.session_state:
    st.session_state['time_unit'] = None
if 'start_time' not in st.session_state:
    st.session_state['start_time'] = '00:00'
if 'end_time' not in st.session_state:
    st.session_state['end_time'] = '23:59'

def generate_time_range(start='00:00', end='23:59', freq='10T'):
    return pd.date_range(start=start, end=end, freq=freq).strftime('%H:%M').tolist()

# 초기 화면
if st.session_state['username'] == '':
    st.title("플래너 만들기")

    username = st.text_input("사용자명을 입력하세요:")
    time_unit_options = {'10분': '10T', '30분': '30T'}
    time_unit = st.selectbox("플래너 단위를 선택하세요:", list(time_unit_options.keys()))
    selected_freq = time_unit_options[time_unit]

    # 시간 범위 생성
    time_options = generate_time_range(end='23:59', freq=selected_freq)
    start_time = st.selectbox("시작 시간을 선택하세요:", time_options, index=min(36, len(time_options)-1))
    end_time = st.selectbox("끝나는 시간을 선택하세요:", time_options, index=min(132, len(time_options)-1))

    if st.button("확인"):
        if username and time_unit and start_time and end_time:
            st.session_state['username'] = username
            st.session_state['time_unit'] = time_unit
            st.session_state['start_time'] = start_time
            st.session_state['end_time'] = end_time
            st.success(f"어서오세요, {username}님!")

if st.session_state['username'] != '' and st.session_state['time_unit']:

    time_freq = '10T' if st.session_state['time_unit'] == '10분' else '30T'
    days_of_week = ['월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일']
    time_slots = generate_time_range(st.session_state['start_time'], st.session_state['end_time'], freq=time_freq)

    if 'weekly_plan' not in st.session_state:
        st.session_state['weekly_plan'] = {day: [''] * len(time_slots) for day in days_of_week}

    st.title(f"{st.session_state['username']}님의 플래너")

    selected_day = st.selectbox("요일을 선택하세요:", days_of_week)
    start_time = st.selectbox("시작 시간을 선택하세요:", time_slots, key='start_time_select')
    end_time = st.selectbox("종료 시간을 선택하세요:", time_slots, key='end_time_select')
    daily_task = st.text_input("계획을 입력하세요:")
    task_color = st.color_picker("계획 색상을 선택하세요:", '#FFFF00')

    if st.button("계획 추가"):
        if daily_task and task_color:
            start_idx = time_slots.index(start_time)
            # end_idx를 포함하지 않도록 수정
            end_idx = time_slots.index(end_time) 
            for idx in range(start_idx, end_idx):
                st.session_state['weekly_plan'][selected_day][idx] = f"<div style='background-color: {task_color};'>{daily_task}</div>"
            st.success(f"{selected_day} {start_time}부터 {end_time}까지의 계획이 추가되었습니다!")
        else:
            st.warning("모든 필드를 입력해주세요.")

    st.subheader("현재 입력된 주간 계획 보기")

    time_rows = "".join([
        f"<tr><td style='padding: 2px 5px; height: 20px;'>{time}</td>" +
        "".join([f"<td style='width: 100px; height: 20px;'>{st.session_state['weekly_plan'][day][i]}</td>"
                 for day in days_of_week]) +
        "</tr>"
        for i, time in enumerate(time_slots)
    ])

    html_table = f"""
    <style>
    table, th, td {{
        border: 1px solid black;
        border-collapse: collapse;
        text-align: center;
    }}
    </style>
    <table style='width:100%;'>
        <thead>
            <tr><th>시간</th>{"".join([f"<th>{day}</th>" for day in days_of_week])}</tr>
        </thead>
        <tbody>
            {time_rows}
        </tbody>
    </table>
    """
    st.markdown(html_table, unsafe_allow_html=True)

    def strip_html(html):
        clean = re.compile('<.*?>')
        return re.sub(clean, '', html)

    # 엑셀 저장시 색깔 코드를 포함하는 함수
    def save_to_excel():
        workbook = Workbook()
        sheet = workbook.active

        # 열 제목 추가
        sheet.append(['시간'] + days_of_week)

        # 데이터 추가
        for i, time in enumerate(time_slots):
            # 시간을 가장 왼쪽 열에 추가
            sheet.cell(row=i + 2, column=1).value = time

            for day_index, day in enumerate(days_of_week):
                cell_value = st.session_state['weekly_plan'][day][i]
                clean_text = strip_html(cell_value)
                cell = sheet.cell(row=i + 2, column=day_index + 2)  # 요일 열은 두번째부터 시작
                cell.value = clean_text

                # HTML 코드에서 색상 정보를 추출하여 셀 색상 지정
                if 'background-color:' in cell_value:
                    color_match = re.search(r'background-color: (#\w+);', cell_value)
                    if color_match:
                        color_code = color_match.group(1).lstrip('#')
                        fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
                        cell.fill = fill

        # 임시 저장소 생성
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            workbook.save(tmp_file.name)
            return tmp_file.name

    # 엑셀 파일 다운로드 버튼
    if st.button("엑셀 파일로 다운로드"):
        excel_path = save_to_excel()
        with open(excel_path, "rb") as file:
            st.download_button("엑셀 파일로 다운로드", data=file, file_name='weekly_plan.xlsx', mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        os.remove(excel_path)  # 다운로드 후 임시 파일 삭제
